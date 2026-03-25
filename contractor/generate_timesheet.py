"""
generate_timesheet.py
---------------------
Standalone timesheet generator. No external files required.
Generates an Excel file matching the TimesheetTemplate.xlsx format:
  - Sheet 1: Instructions (ditto copy)
  - Sheet 2: Payroll Timesheet Info (your data)

Requirements: pip install openpyxl
Usage:        python generate_timesheet.py
"""

# ── SECTION 1: Imports ────────────────────────────────────────────────────────
import os
import openpyxl
from openpyxl.styles import Font
from datetime import date, timedelta

# ── SECTION 2: Hardcoded Data ─────────────────────────────────────────────────

EMPLOYEE_IDS = [
    "4828", "4690", "4422", "5005", "4691", "4831", "3636", "4846",
    "4730", "1143", "4362", "4733", "5024", "4913", "4914", "4963",
    "4684", "3685", "4978", "4917", "5043", "4772", "4990", "4954",
    "4700", "4767", "5017", "4017", "FZ-0041", "4744", "4942", "4222",
    "FZ-0042", "4681", "5032", "4524", "FZ-0044", "FZ-0043",
]  # 38 total

PROJECT_CODES = [
    "20-00632", "22-00656", "23-00608", "23-00616", "23-00618", "23-00622",
    "23-00626", "23-00627", "24-00603", "24-00606", "24-00607", "24-00612",
    "24-00613", "24-00615", "24-00616", "24-00617", "24-00618", "24-00623",
    "24-00626", "24-00627", "24-00629", "24-00630", "24-00631", "24-00632",
    "25-00602", "25-00603", "25-00604", "25-00605", "25-00607", "25-00609",
    "25-00610", "25-00612", "25-00613", "25-00614", "25-00617", "25-00620",
    "25-00621", "25-00622", "25-00623", "25-00624", "25-00625", "25-00626",
    "25-00627", "25-00628", "25-00629", "25-00630", "25-00631", "25-00632",
    "25-00633", "25-00634", "25-00635", "25-00637", "25-00638", "AOH-260001",
    "AOH-260003", "AOH-260004", "AOH-260005", "AOH-260010", "AOH-260012",
    "AOH-260014", "DLESAFTOOL", "OMNISAFTOL", "000000OTH", "SAFETYTOOL",
    "25-00608",
]  # 65 total

# Exact ditto copy of TimesheetTemplate.xlsx Sheet 1 — Instructions
INSTRUCTIONS_DATA = [
    ("Column Name",              "Column Description",                        "Mandatory"),
    ("Emp ID",                   "",                                           "True"),
    ("Date Worked",              "yyyy-MM-dd format (e.g. 2023-12-25)",        "True"),
    ("Earnings Code",            "RG, OT, DT, VL, SK, BP",                    "True"),
    ("Hours",                    "HH:MM format (e.g. 8:00)",                   "True"),
    ("Project",                  "If no project then link to ProjectX",        ""),
    ("Project Description",      "",                                           ""),
    ("Task",                     "",                                           ""),
    ("Task Description",         "",                                           ""),
    ("Cost Code",                "",                                           ""),
    ("Cost Type",                "",                                           ""),
    ("Labor Item",               "",                                           ""),
    ("Labor item Description",   "",                                           ""),
    ("Union",                    "",                                           ""),
    ("Card Type",                "",                                           ""),
    ("Work Status",              "",                                           ""),
    ("Shift",                    "",                                           ""),
    ("Special Pay",              "",                                           ""),
    ("Sub",                      "",                                           ""),
    ("Direct",                   "",                                           ""),
    ("CertExcl",                 "",                                           ""),
    ("Flag2",                    "",                                           ""),
    ("Flag1",                    "",                                           ""),
    ("End Date",                 "yyyy-MM-dd format (e.g. 2023-12-25)",        ""),
]  # 24 rows (row 1 is the header itself)

TIMESHEET_HEADERS = [
    "Emp ID", "Date Worked", "Earnings Code", "Hours", "Project",
    "Project Description", "Task", "Task Description", "Cost Code",
    "Cost Type", "Labor Item", "Labor item Description", "Union",
    "Card Type", "Work Status", "Shift", "Special Pay", "Sub",
    "Direct", "CertExcl", "Flag2", "Flag1", "End Date",
]  # 23 columns (A through W)

# ── SECTION 3: Configuration ──────────────────────────────────────────────────

DEFAULT_HOURS = {
    "RG": "8:00",
    "OT": "10:00",
    "DT": "14:00",
}

VALID_EARNINGS_CODES = ["RG", "OT", "DT"]
MAX_EMPLOYEES = 38
MAX_DAYS = 7

# ── SECTION 4: Helpers ────────────────────────────────────────────────────────

def get_monday():
    today = date.today()
    return today - timedelta(days=today.weekday())


# ── SECTION 5: Prompts ────────────────────────────────────────────────────────

def prompt_num_employees():
    while True:
        raw = input(f"How many employees? (1–{MAX_EMPLOYEES}, default {MAX_EMPLOYEES}): ").strip()
        if raw == "":
            return MAX_EMPLOYEES
        try:
            n = int(raw)
        except ValueError:
            print("  Please enter a whole number.")
            continue
        if 1 <= n <= MAX_EMPLOYEES:
            return n
        print(f"  Must be between 1 and {MAX_EMPLOYEES}.")


def prompt_num_days():
    while True:
        raw = input(f"How many days per employee? (1–{MAX_DAYS}): ").strip()
        try:
            n = int(raw)
        except ValueError:
            print("  Please enter a whole number.")
            continue
        if 1 <= n <= MAX_DAYS:
            return n
        print(f"  Must be between 1 and {MAX_DAYS}.")


def prompt_earnings_code():
    code_map = {"1": "RG", "2": "OT", "3": "DT"}
    while True:
        raw = input("Earnings code? [1=RG / 2=OT / 3=DT]: ").strip()
        if raw in code_map:
            return code_map[raw]
        print("  Invalid. Enter 1, 2, or 3.")


def prompt_output_filename():
    raw = input("Output filename (without .xlsx, default: Timesheet_Output): ").strip()
    if not raw:
        return "Timesheet_Output"
    if raw.lower().endswith(".xlsx"):
        raw = raw[:-5]
    return raw

# ── SECTION 6: Sheet Builders ─────────────────────────────────────────────────

def build_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions", 0)
    bold = Font(bold=True)
    for i, (col_name, col_desc, mandatory) in enumerate(INSTRUCTIONS_DATA, start=1):
        ws.cell(row=i, column=1, value=col_name)
        ws.cell(row=i, column=2, value=col_desc)
        ws.cell(row=i, column=3, value=mandatory)
        if i == 1:
            ws.cell(row=i, column=1).font = bold
            ws.cell(row=i, column=2).font = bold
            ws.cell(row=i, column=3).font = bold
    # Column widths for readability
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12


def build_timesheet_sheet(wb, emp_ids, dates, code, hours_str):
    ws = wb.create_sheet("Payroll Timesheet Info", 1)
    bold = Font(bold=True)

    # Header row
    for col_idx, header in enumerate(TIMESHEET_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Data rows
    row_num = 2
    proj_idx = 0
    for emp_id in emp_ids:
        for day_val in dates:
            project = PROJECT_CODES[proj_idx % len(PROJECT_CODES)]
            ws.cell(row=row_num, column=1, value=str(emp_id))
            date_cell = ws.cell(row=row_num, column=2, value=day_val)
            date_cell.number_format = "yyyy-MM-dd"
            ws.cell(row=row_num, column=3, value=code)
            ws.cell(row=row_num, column=4, value=hours_str)
            ws.cell(row=row_num, column=5, value=project)
            # Columns 6–23 left blank
            row_num += 1
            proj_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14

# ── SECTION 7: Entry Point ────────────────────────────────────────────────────

def main():
    print("=" * 45)
    print("       TIMESHEET GENERATOR")
    print("=" * 45)

    num_emps = prompt_num_employees()
    num_days = prompt_num_days()
    code     = prompt_earnings_code()
    hours_str = DEFAULT_HOURS[code]

    monday = get_monday()
    dates  = [monday + timedelta(days=i) for i in range(num_days)]  # date objects
    emp_ids = EMPLOYEE_IDS[:num_emps]

    print(f"\n  Week starts : {dates[0].strftime('%Y-%m-%d')} (Monday)")
    if num_days > 1:
        print(f"  Week ends   : {dates[-1].strftime('%Y-%m-%d')}")

    out_name = prompt_output_filename()
    full_path = out_name + ".xlsx"

    # Overwrite guard
    if os.path.exists(full_path):
        confirm = input(f"\n  '{full_path}' already exists. Overwrite? (y/n): ").strip().lower()
        if confirm != "y":
            print("  Cancelled.")
            return

    # Build workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    build_instructions_sheet(wb)
    build_timesheet_sheet(wb, emp_ids, dates, code, hours_str)

    try:
        wb.save(full_path)
    except PermissionError:
        print(f"\n  ERROR: Cannot save '{full_path}' — is the file open in Excel?")
        return
    except Exception as e:
        print(f"\n  ERROR: {e}")
        return

    total_rows = num_emps * num_days

    print()
    print(f"  Created : {full_path}")
    print(f"  Employees : {num_emps}")
    print(f"  Days      : {num_days}")
    print(f"  Rows      : {num_emps} x {num_days} = {total_rows}")
    print(f"  Code      : {code}  |  Hours: {hours_str}")
    print()

    # Console preview (first 5 rows)
    print("  Preview (first 5 data rows):")
    print(f"  {'Emp ID':<10} {'Date':<12} {'Code':<6} {'Hours':<6} {'Project'}")
    print(f"  {'-'*10} {'-'*12} {'-'*6} {'-'*6} {'-'*14}")
    proj_idx = 0
    for emp_id in emp_ids:
        for day_val in dates:
            project = PROJECT_CODES[proj_idx % len(PROJECT_CODES)]
            print(f"  {str(emp_id):<10} {day_val.strftime('%Y-%m-%d'):<12} {code:<6} {hours_str:<6} {project}")
            proj_idx += 1
            if proj_idx >= 5:
                break
        if proj_idx >= 5:
            break
    if total_rows > 5:
        print(f"  ... ({total_rows - 5} more rows)")
    print()


if __name__ == "__main__":
    main()
